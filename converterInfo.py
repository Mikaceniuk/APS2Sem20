from openpyxl import load_workbook

def converterExcelParaTxt(nomeExcel, nomeCabecalho):
    print('Convertendo {}'.format(nomeExcel))

    path = '{}.xlsx'.format(nomeExcel)
    wb_obj = load_workbook(path)
    f = open('{}.txt'.format(nomeExcel), 'w+')
    sheet_obj = wb_obj.active

    cabecalhos = []
    valores = []

    #Pega o cabeçalho e valor de cada linha
    for row in range(sheet_obj.max_row + 1):
        valorLinha = {}
        for column in range(1, sheet_obj.max_column + 1):
            if(row > 0):
                valorCelula = sheet_obj.cell(row = row, column = column).value
                if(row == 1):
                    cabecalhos.append(valorCelula)
                    if(column != sheet_obj.max_column):
                        f.writelines('{} | '.format(valorCelula))
                    else:
                        f.writelines('{} \n'.format(valorCelula))
                elif(row > 1):
                    valorLinha['{}'.format(cabecalhos[column - 1])] = valorCelula                    
        if(row > 1):
            valores.append(valorLinha)

    #Ordenar
    indexCab = cabecalhos.index(nomeCabecalho)
    for i in range(len(valores)):
        if(i > 0):
            vlrLinhaAtual = valores[i]
            vlrLinhaAnterior = valores[i - 1]

            while int(vlrLinhaAnterior.get(nomeCabecalho)) > int(vlrLinhaAtual.get(nomeCabecalho)):
                indexLinhaAnt = valores.index(vlrLinhaAtual)
                valores[valores.index(vlrLinhaAnterior)] = valores[valores.index(vlrLinhaAtual)]
                valores[indexLinhaAnt] = vlrLinhaAnterior

                nvIndex = valores.index(vlrLinhaAtual)
                if(nvIndex == 0):
                    i = 0
                    break
                vlrLinhaAtual = valores[nvIndex]
                vlrLinhaAnterior = valores[nvIndex - 1]

    #Escreve
    for i in range(len(valores)):
        for x in cabecalhos:
            if(cabecalhos.index(x) != len(cabecalhos) - 1):
                f.writelines('{} | '.format(valores[i][x]))
            else:
                f.writelines('{} \n'.format(valores[i][x]))

    f.close()

    print('Convertido {} com sucesso!'.format(nomeExcel))